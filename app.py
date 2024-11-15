from flask import Flask, render_template, redirect, request, flash
import openpyxl

app = Flask(__name__)
app.secret_key = 'senha12345678'

@app.route('/')
def home():
    # Exibe a página inicial com o menu de pizzas e formulários de cadastro
    return render_template('index.html')

@app.route('/index', methods=['POST'])
def cadastrar():
    # Dados do pedido do cliente
    nome_cliente = request.form.get('firstname') + ' ' + request.form.get('lastname')
    email = request.form.get('email')
    numero = request.form.get('number')
    pizza_escolhida = request.form.get('pizza')  # A pizza escolhida pelo cliente
    tamanho_pizza = request.form.get('size')  # Tamanho da pizza
    data_entrega = request.form.get('data')  # Data da entrega

    # Verifica se o cliente já fez um pedido anteriormente
    if buscar(nome_cliente):
        flash(f'O cliente {nome_cliente} já possui um pedido registrado!', 'error')
    else:
        # Salva os dados do pedido
        save(nome_cliente, email, numero, pizza_escolhida, tamanho_pizza, data_entrega)
        flash(f'Pedido registrado com sucesso! Sua pizza {pizza_escolhida} será entregue em breve.', 'success')
    
    return redirect('/')

@app.route('/buscar', methods=['POST'])
def buscar_pedido():
    # Busca o pedido do cliente
    nome_cliente = request.form.get('nome')

    if buscar(nome_cliente):
        flash(f'Pedido encontrado para o cliente: {nome_cliente}.', 'success')
    else:
        flash(f'O cliente {nome_cliente} não possui nenhum pedido registrado.', 'error')
        
    return redirect('/')

def save(nome_cliente, email, numero, pizza_escolhida, tamanho_pizza, data_entrega):
    try:
        # Carrega o arquivo Excel onde os pedidos são salvos
        book = openpyxl.load_workbook('pedidos_pizza.xlsx')
    except FileNotFoundError:
        # Cria um novo arquivo Excel caso não exista
        book = openpyxl.Workbook()

    if 'Pedidos' not in book.sheetnames:
        book.create_sheet('Pedidos')
    
    banco_pg = book['Pedidos']
    next_row = 1
    
    # Encontra a próxima linha livre para salvar o pedido
    while banco_pg.cell(row=next_row, column=2).value is not None:
        next_row += 1

    # Preenche as informações do pedido na próxima linha disponível
    banco_pg.cell(row=next_row, column=2, value=nome_cliente)
    banco_pg.cell(row=next_row, column=3, value=email)
    banco_pg.cell(row=next_row, column=4, value=numero)
    banco_pg.cell(row=next_row, column=5, value=pizza_escolhida)
    banco_pg.cell(row=next_row, column=6, value=tamanho_pizza)
    banco_pg.cell(row=next_row, column=7, value=data_entrega)
    
    # Salva o arquivo de pedidos
    book.save('pedidos_pizza.xlsx')

def buscar(nome_cliente):
    # Verifica se o pedido do cliente existe no arquivo Excel
    book = openpyxl.load_workbook('pedidos_pizza.xlsx')
    banco_pg = book['Pedidos']
    
    # Itera sobre todas as linhas para encontrar o nome do cliente
    for row in banco_pg.iter_rows(values_only=True):
        if row[1] == nome_cliente:
            return True  
    return False 

if __name__ == "__main__":
    app.run(debug=True)
